from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import pdfplumber
import pandas as pd
import re
import io
import json
import base64
from collections import defaultdict

# ---------- CONSTANTS & HELPERS ----------
GRADE_POINTS = {
    'S': 10, 'A+': 9, 'A': 8.5, 'B+': 8, 'B': 7.5, 'C+': 7,
    'C': 6.5, 'D': 6, 'P': 5.5, 'PASS': 5.5,
    'F': 0, 'FE': 0, 'I': 0, 'ABSENT': 0, 'WITHHELD': 0
}

# register number regex: match PKD or LPKD prefix with two‑digit year,
# two letters dept code and three digits. used for filtering stray rows.
REG_NO_PATTERN = re.compile(r'(L?PKD\d{2}[A-Z]{2}\d{3})')
COURSE_GRADE_PATTERN = re.compile(r'([A-Z]{3,}\d{3})\s*\(([^)]+)\)')


def detect_metadata(text: str):
    """Return (semester, scheme, exam_name) detected from header text."""
    sem_match = re.search(r'\b(S[1-8])\b|SEMESTER\s+([IVX]+|[1-8])', text, re.IGNORECASE)
    semester = "S1"
    if sem_match:
        if sem_match.group(1):
            semester = sem_match.group(1).upper()
        else:
            roman = {"I":"S1","II":"S2","III":"S3","IV":"S4","V":"S5","VI":"S6","VII":"S7","VIII":"S8"}
            semester = roman.get(sem_match.group(2).upper(), semester)

    scheme = "2024" if "2024" in text else "2019"
    exam_name = "B.Tech Degree Examination" if "B.Tech" in text else "University Examination"
    return semester, scheme, exam_name


def get_course_credits(code: str, lookup: dict) -> int:
    """Look up credits for a given course code; supports X wildcard patterns."""
    clean = code.replace(" ", "")
    if clean in lookup:
        return lookup[clean]
    for pattern, val in lookup.items():
        if 'X' in pattern or 'N' in pattern[-1:]:
            regex = "^" + pattern.replace("X", "[A-Z0-9]").replace("N", "[A-Z0-9]") + "$"
            if re.match(regex, clean):
                return val
    return 0


app = FastAPI(title="KTU Result Analyzer API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def extract_and_analyze(pdf_bytes: bytes):
    departments_data = {}
    stats_data = []
    total_students = 0
    total_passed = 0
    missing_credit_courses = defaultdict(int)  # count of lookups that returned zero

    # will collect entire PDF text for metadata detection and SGPA logic
    full_text = ""
    
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        department = "Unknown"
        
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            full_text += text + "\n"
                
            lines = text.split('\n')
            current_student = None
            
            for line in lines:
                # 1. Match Department - Handle various KTU formats (e.g. Branch, Full Time, Generated on, etc.)
                # Often KTU headers look like: "INFORMATION TECHNOLOGY[Full Time] (Generated on ...)"
                # Or just the branch name before the date
                dept_match = re.match(r'^(.*?)(?:\[Full Time\]|\[Part Time\]|\(Generated on)', line)
                if dept_match and len(dept_match.group(1).strip()) > 5:
                    # Ignore lines that are definitely not departments
                    potential_dept = dept_match.group(1).strip()
                    if not potential_dept.startswith('APJ ABDUL KALAM') and not potential_dept.startswith('Exam Centre:'):
                        department = potential_dept
                        if department not in departments_data:
                            departments_data[department] = []
                        current_student = None
                        continue
                
                # Match Course string pattern globally: Any alphanumeric/hyphen/underscore string followed by (Grade)
                results_matches = re.findall(r'([A-Za-z0-9\-\_]+)\(([\w\+]+)\)', line)
                
                # 2. Match New Student Line
                student_match = re.match(r'^([A-Z0-9]{9,15})\s+(.*)', line)
                if student_match:
                    student_id = student_match.group(1)
                    
                    # Save previous student if exists
                    if current_student:
                        if department not in departments_data:
                            departments_data[department] = []
                        departments_data[department].append(current_student)
                        
                    current_student = {'Student ID': student_id}
                    
                    # Add courses on this line
                    for course, grade in results_matches:
                        current_student[course] = grade
                        
                # 3. Match continuation of student courses on next line
                elif current_student and len(results_matches) > 0:
                    for course, grade in results_matches:
                        current_student[course] = grade
                        
                # 4. If we hit empty line or something else, close the student block
                elif current_student and not line.strip():
                    if department not in departments_data:
                        departments_data[department] = []
                    departments_data[department].append(current_student)
                    current_student = None
                    
            # Add final student on page if exists
            if current_student:
                if department not in departments_data:
                    departments_data[department] = []
                departments_data[department].append(current_student)
                current_student = None

        # metadata detection & credit lookup
        detected_semester, detected_scheme, exam_name = detect_metadata(full_text)
        credits_file = f"credits_{detected_scheme}.json"
        # ensure we load from backend directory regardless of working directory
        import os
        base_dir = os.path.dirname(__file__)
        credits_path = os.path.join(base_dir, credits_file)
        try:
            with open(credits_path, "r") as f:
                credit_data = json.load(f)
        except FileNotFoundError:
            # if file missing, leave empty and later SGPA will be zero
            credit_data = {}
            print(f"WARNING: credit file not found at {credits_path}")

        # helper to extract entry year from register number
        def extract_year(reg: str):
            m = re.match(r'^L?PKD(\d{2})', reg, re.IGNORECASE)
            if m:
                return m.group(1)
            # fallback: look for any two-digit year at start
            m2 = re.match(r'.*?(\d{2})', reg)
            return m2.group(1) if m2 else None

        # separate regular vs supplementary students by year mismatch
        reg_dept = defaultdict(list)
        supp_dept = defaultdict(list)
        for dept, students in departments_data.items():
            # determine main year for this department - prefer the most frequent,
            # break ties by choosing the earliest year (lower number).
            year_counts = defaultdict(int)
            for s in students:
                yr = extract_year(s.get('Student ID', ''))
                if yr:
                    year_counts[yr] += 1
            main_year = None
            if year_counts:
                max_count = max(year_counts.values())
                candidates = [y for y, c in year_counts.items() if c == max_count]
                main_year = min(candidates)

            for s in students:
                yr = extract_year(s.get('Student ID', ''))
                # entries without a recognizable year go to supplementary
                if main_year and yr and yr != main_year:
                    supp_dept[dept].append(s)
                elif main_year and not yr:
                    supp_dept[dept].append(s)
                else:
                    reg_dept[dept].append(s)

        # remove any stray rows that lack a valid KTU register number
        # such as header lines like ELECTRICAL, ESSENTIALS etc. (lateral
        # students begin with LPKD so the regex now accepts an optional L)
        cleaned = {}
        for dept, studs in reg_dept.items():
            valid = [s for s in studs if REG_NO_PATTERN.match(s.get('Student ID',''))]
            if valid:
                cleaned[dept] = valid
        reg_dept = cleaned

        # also scrub supplementary bucket to avoid bogus rows lingering
        cleaned_supp = {}
        for dept, studs in supp_dept.items():
            valid = [s for s in studs if REG_NO_PATTERN.match(s.get('Student ID',''))]
            if valid:
                cleaned_supp[dept] = valid
        supp_dept = cleaned_supp

        # after cleaning, regular buckets are used for analysis; supplementary
        # students are kept in supp_dept to be written to a separate workbook if
        # any exist.
        dept_buckets = reg_dept
        departments_data = reg_dept

        # metadata and credit loading continues...
        # build semester_totals based on scheme
        if detected_scheme == "2024":
            semester_totals = credit_data.get("semester_total_credits", {})
        else:
            # 2019 structure: departments -> semesters each with total_credit
            semester_totals = {}
            for dept in credit_data.get("departments", []):
                for sem in dept.get("semesters", []):
                    key = f"S{sem.get('semester')}"
                    total = sem.get("total_credit", 0)
                    if key in semester_totals:
                        semester_totals[key] = max(semester_totals[key], total)
                    else:
                        semester_totals[key] = total
        semester_key = detected_semester
        if semester_key not in semester_totals:
            semester_key = detected_semester.replace("S", "")

        # build credit lookup: handles both JSON formats
        credit_lookup = {}
        if detected_scheme == "2024":
            credit_lookup = {
                c["code"].replace(" ", ""): c["credits"]
                for d in credit_data.get("curricula", [])
                for s in d.get("semesters", [])
                for c in s.get("courses", [])
            }
        else:
            # 2019 scheme: iterate departments and their semester courses
            for dept in credit_data.get("departments", []):
                for sem in dept.get("semesters", []):
                    for c in sem.get("courses", []):
                        code = c.get("course_code", "").replace(" ", "")
                        credit_lookup[code] = c.get("credit", 0)
        if not credit_lookup:
            print(f"WARNING: credit lookup is empty (scheme {detected_scheme}). check {credits_path}")

        # --- Handle Multiple Electives Selection ---
        # In some schemes/semesters (like S6 2019), multiple electives might be listed.
        # If a student passes at least one elective in a group, the failed ones in that group
        # are ignored and do not affect their pass/fail status or total credits.
        ELECTIVE_GROUPS = [
            {"ECT352", "ECT362", "ECT372", "ECT374", "ECT376", "ECT378", "ECT382", "ECT384"},
            {"CST342", "CST352", "CSL362", "CST362", "CST372"},
            {"EET312", "EET322", "EET332"},
            {"CET312", "CET322", "CET332", "CET342", "CET352", "CET362"}
        ]
        
        all_students_lists = list(reg_dept.values()) + list(supp_dept.values())
        for students in all_students_lists:
            for student in students:
                for group in ELECTIVE_GROUPS:
                    courses_in_grp = [c for c in group if c in student]
                    if len(courses_in_grp) > 1:
                        # Check if passed at least one
                        passed_any = any(str(student[c]).upper() not in ['F', 'FE', 'ABSENT', 'DEBARRED', 'I', 'WITHHELD'] for c in courses_in_grp)
                        if passed_any:
                            # Remove the failing ones so they don't count as arrear or fail the student
                            for c in courses_in_grp:
                                if str(student[c]).upper() in ['F', 'FE', 'ABSENT', 'DEBARRED', 'I', 'WITHHELD']:
                                    del student[c]

        # Calculate Results Post-Extraction
        for dept, students in departments_data.items():
            for student in students:
                passed_all = True
                total_weighted = 0
                total_creds = 0
                for course, grade in student.items():
                    if course in ['Student ID', 'Result', 'SGPA']:
                        continue
                    if str(grade).upper() in ['F', 'FE', 'ABSENT', 'DEBARRED', 'I', 'WITHHELD']:
                        passed_all = False
                    gp = GRADE_POINTS.get(grade.upper(), 0)
                    creds = get_course_credits(course, credit_lookup)
                    if creds == 0:
                        missing_credit_courses[course] += 1
                    total_weighted += creds * gp
                    # 2019: include ALL courses in denominator (per official formula)
                    # 2024: only count passed courses for the denominator
                    if detected_scheme == "2019":
                        total_creds += creds
                    else:
                        if grade not in ['F', 'FE', 'Absent', 'Debarred', 'I']:
                            total_creds += creds
                # account for special 2024 S2 extra credit/activity
                if detected_scheme == "2024" and detected_semester == "S2":
                    total_creds += 1
                    total_weighted += 1 * 5.5
                # use different denominators based on scheme
                if detected_scheme == "2019":
                    # 2019 formula: SGPA = Σ(Ci×GPi)/ΣCi (sum of all course credits)
                    sgpa_denom = total_creds if total_creds > 0 else 1
                else:
                    # 2024: use official semester total
                    sgpa_denom = (
                        24 if detected_semester == "S2"
                        else semester_totals.get(semester_key, 21)
                    )
                sgpa = round(total_weighted / sgpa_denom, 2) if sgpa_denom else 0
                student['SGPA'] = sgpa
                student['Result'] = 'PASS' if passed_all else 'FAIL'

    if not departments_data:
        raise ValueError("Could not extract any results from the PDF.")

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Calculate Arrears for all students
    for dept, students in list(reg_dept.items()) + list(supp_dept.items()):
        for student in students:
            arrear_count = 0
            for course, grade in student.items():
                if course in ['Student ID', 'Result', 'SGPA', 'Arrear']:
                    continue
                if str(grade).upper() in ['F', 'FE', 'ABSENT', 'DEBARRED', 'I', 'WITHHELD']:
                    arrear_count += 1
            student['Arrear'] = arrear_count

    # Define Styles
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    black_font = Font(color="000000", bold=True, size=11)
    red_font_bold = Font(color="FF0000", bold=True)
    red_font_italic = Font(color="FF0000", italic=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    inst_match = re.search(r'Exam\s*Centre:\s*(.+)', full_text, re.IGNORECASE)
    if inst_match:
        college_name = "Exam Centre: " + inst_match.group(1).strip()
    else:
        college_name = full_text.split('\n')[0].strip() if full_text else "INSTITUTION"
        if "APJ ABDUL KALAM" in college_name and len(full_text.split('\n')) > 1:
            college_name = "Exam Centre: " + full_text.split('\n')[1].strip()

    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        for dept, students in reg_dept.items():
            if not students:
                continue
                
            df = pd.DataFrame(students)
            
            # Organize columns
            base_cols = ['Student ID']
            end_cols = ['SGPA', 'Arrear']
            # Find all courses
            courses = [c for c in df.columns if c not in base_cols + end_cols + ['Result']]
            # Fill missing grades with '-'
            for c in courses:
                if c not in df:
                    df[c] = '-'
                else:
                    df[c] = df[c].fillna('-')
            
            if 'SGPA' not in df: df['SGPA'] = 'N/A'
            if 'Arrear' not in df: df['Arrear'] = 0
            
            # Format SGPA
            df['SGPA'] = df['SGPA'].apply(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) and x != 0 else 'N/A')
            
            # Analytics for Regular Students only
            dept_total = len(df)
            dept_passed = len(df[df['Result'] == 'PASS']) if not df.empty and 'Result' in df else 0
            dept_failed = dept_total - dept_passed

            cols = base_cols + courses + end_cols
            df = df[cols]
            
            # Sort by Register Number
            df = df.sort_values(by=['Student ID']).reset_index(drop=True)
            dept_pass_percentage = round((dept_passed / dept_total) * 100, 2) if dept_total > 0 else 0
            
            total_students += dept_total
            total_passed += dept_passed
            
            # Write to Excel, starting from row 8 (row 7 is empty)
            df.to_excel(writer, sheet_name=dept[:31], index=False, startrow=7)
            
            workbook = writer.book
            worksheet = writer.sheets[dept[:31]]
            
            max_col = len(cols)
            max_col_letter = get_column_letter(max_col)
            
            # --- Headers (Rows 1-6) ---
            worksheet.merge_cells(f"A1:{max_col_letter}1")
            cell = worksheet["A1"]
            cell.value = "KTU RESULT ANALYSER"
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True, size=14)
            cell.alignment = center_align
            
            worksheet.merge_cells(f"A2:{max_col_letter}2")
            cell = worksheet["A2"]
            cell.value = f"{exam_name} ({detected_scheme} Scheme) ({detected_semester} Result)"
            cell.fill = green_fill
            cell.font = black_font
            cell.alignment = center_align
            
            worksheet.merge_cells(f"A3:{max_col_letter}3")
            cell = worksheet["A3"]
            cell.value = f"Department: {dept}"
            cell.fill = yellow_fill
            cell.font = black_font
            cell.alignment = center_align
            
            worksheet.merge_cells(f"A4:{max_col_letter}4")
            cell = worksheet["A4"]
            cell.value = college_name
            cell.fill = yellow_fill
            cell.font = black_font
            cell.alignment = center_align
            
            worksheet.merge_cells(f"A5:{max_col_letter}5")
            cell = worksheet["A5"]
            cell.value = f"Total Regular Students: {len(students)}"
            cell.fill = yellow_fill
            cell.font = black_font
            cell.alignment = center_align
            
            worksheet.merge_cells(f"A6:{max_col_letter}6")
            cell = worksheet["A6"]
            cell.value = "Note: Table shows REGULAR students only."
            cell.font = red_font_italic
            cell.alignment = center_align
            
            # Format Table Headers (Row 8)
            for col_num in range(1, max_col + 1):
                cell = worksheet.cell(row=8, column=col_num)
                if cell.value == "Student ID": cell.value = "Register No"
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                
            # Format Data Rows (Row 9 onwards)
            fail_grades = ['F', 'FE', 'ABSENT', 'DEBARRED', 'I', 'WITHHELD']
            for row_num in range(9, 9 + len(df)):
                for col_num in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.alignment = center_align
                    cell.border = thin_border
                    val = str(cell.value).strip().upper()
                    if val in fail_grades:
                        cell.font = red_font_bold
                    # Highlight Arrear > 0 in red
                    if col_num == max_col and str(cell.value) != '0' and str(cell.value) != 'N/A':
                        cell.font = red_font_bold
                    
                    if col_num == 1:
                        cell.font = Font(color="00B050", bold=True)
                            
            # Auto-adjust column widths
            for col_num in range(1, max_col + 1):
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = 12
            worksheet.column_dimensions['A'].width = 18

            # --- PERFORMANCE ANALYSIS SECTION ---
            start_row = 9 + len(df) + 2
            
            worksheet.merge_cells(f"A{start_row}:{max_col_letter}{start_row}")
            cell = worksheet[f"A{start_row}"]
            cell.value = "PERFORMANCE ANALYSIS - REGULAR STUDENTS ONLY"
            cell.fill = orange_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
            worksheet[f"A{start_row+1}"] = "Pass Percentage"
            worksheet[f"B{start_row+1}"] = f"{dept_pass_percentage}%"
            worksheet[f"A{start_row+2}"] = "Total Regular"
            worksheet[f"B{start_row+2}"] = dept_total
            worksheet[f"A{start_row+3}"] = "Total Passed"
            worksheet[f"B{start_row+3}"] = dept_passed
            worksheet[f"A{start_row+4}"] = "Total Failed"
            worksheet[f"B{start_row+4}"] = dept_failed
            
            for r in range(start_row+1, start_row+5):
                for c in ["A", "B"]:
                    cell = worksheet[f"{c}{r}"]
                    cell.border = thin_border
                    cell.font = black_font
                    if c == "A": cell.fill = header_fill; cell.font = header_font
                    cell.alignment = center_align

            # --- SUBJECT-WISE ANALYSIS SECTION ---
            subj_start = start_row + 6
            worksheet.merge_cells(f"A{subj_start}:{max_col_letter}{subj_start}")
            cell = worksheet[f"A{subj_start}"]
            cell.value = "SUBJECT-WISE ANALYSIS - REGULAR STUDENTS ONLY"
            cell.fill = orange_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
            subj_headers = ["SubCode", "Pass%", "Pass", "Fail", "S", "A+", "A", "B+", "B", "C+", "C", "D", "P", "F", "FE"]
            for i, h in enumerate(subj_headers):
                cell = worksheet.cell(row=subj_start+1, column=i+1)
                cell.value = h
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                
            subject_stats = []
            if not df.empty:
                for idx, course in enumerate(courses):
                    if course in df:
                        course_grades = df[course].dropna()
                        course_grades_upper = course_grades.astype(str).str.upper()
                        fails = course_grades_upper.isin(['F', 'FE', 'ABSENT', 'DEBARRED', 'I', 'WITHHELD']).sum()
                        passes = len(course_grades_upper[~course_grades_upper.isin(['F', 'FE', 'ABSENT', 'DEBARRED', 'I', 'WITHHELD', '-'])])
                        total_valid = passes + fails
                        pass_pct = round((passes / total_valid) * 100, 2) if total_valid > 0 else 0
                        
                        subject_stats.append({
                            "subject": course,
                            "passed": int(passes),
                            "failed": int(fails)
                        })
                        
                        r = subj_start + 2 + idx
                        worksheet.cell(row=r, column=1, value=course)
                        worksheet.cell(row=r, column=2, value=f"{pass_pct}%")
                        worksheet.cell(row=r, column=3, value=passes)
                        worksheet.cell(row=r, column=4, value=fails)
                        
                        grades = ['S', 'A+', 'A', 'B+', 'B', 'C+', 'C', 'D', 'P', 'F', 'FE']
                        for g_idx, g in enumerate(grades):
                            count = course_grades.isin([g]).sum()
                            worksheet.cell(row=r, column=5+g_idx, value=count)
                            
                        # Format row
                        for c_idx in range(1, len(subj_headers)+1):
                            cell = worksheet.cell(row=r, column=c_idx)
                            cell.alignment = center_align
                            cell.border = thin_border

            stats_data.append({
                "name": dept.title(),
                "total": dept_total,
                "passed": dept_passed,
                "failed": dept_failed,
                "passPercentage": dept_pass_percentage,
                "subjectStats": subject_stats
            })

            # --- TOP 10 STUDENTS SECTION ---
            top_start = subj_start + 2 + len(courses) + 2
            worksheet.merge_cells(f"A{top_start}:C{top_start}")
            cell = worksheet[f"A{top_start}"]
            cell.value = "TOP 10 STUDENTS (REGULAR)"
            cell.fill = orange_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
            worksheet[f"A{top_start+1}"] = "Rank"
            worksheet[f"B{top_start+1}"] = "Register No"
            worksheet[f"C{top_start+1}"] = "SGPA"
            for c in ["A", "B", "C"]:
                cell = worksheet[f"{c}{top_start+1}"]
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
                
            # Filter valid SGPA and sort
            if not df.empty and 'SGPA' in df:
                valid_sgpa = df[pd.to_numeric(df['SGPA'], errors='coerce').notnull()].copy()
                valid_sgpa['numeric_sgpa'] = pd.to_numeric(valid_sgpa['SGPA'])
                top_10 = valid_sgpa.sort_values(by='numeric_sgpa', ascending=False).head(10)
                
                curr_rank = 1
                curr_row = top_start + 2
                prev_sgpa = None
                
                for _, student in top_10.iterrows():
                    sgpa = student['numeric_sgpa']
                    if prev_sgpa is not None and sgpa < prev_sgpa:
                        curr_rank += 1
                    worksheet.cell(row=curr_row, column=1, value=curr_rank)
                    worksheet.cell(row=curr_row, column=2, value=student['Student ID'])
                    worksheet.cell(row=curr_row, column=3, value=f"{sgpa:.2f}")
                    
                    for c_idx in range(1, 4):
                        cell = worksheet.cell(row=curr_row, column=c_idx)
                        cell.alignment = center_align
                        cell.border = thin_border
                    
                    prev_sgpa = sgpa
                    curr_row += 1

    excel_buffer.seek(0)
    excel_base64 = base64.b64encode(excel_buffer.read()).decode('utf-8')

    # ----- SUPPLEMENTARY WORKBOOK -----
    supp_base64 = None
    supp_total = sum(len(v) for v in supp_dept.values())
    
    if supp_total > 0:
        supp_buffer = io.BytesIO()
        with pd.ExcelWriter(supp_buffer, engine='openpyxl') as writer:
            for dept, students in supp_dept.items():
                if not students:
                    continue
                df = pd.DataFrame(students)
                
                base_cols = ['Student ID']
                end_cols = ['SGPA', 'Arrear']
                courses = [c for c in df.columns if c not in base_cols + end_cols + ['Result']]
                for c in courses:
                    if c not in df: df[c] = '-'
                    else: df[c] = df[c].fillna('-')
                
                if 'SGPA' not in df: df['SGPA'] = 'N/A'
                if 'Arrear' not in df: df['Arrear'] = 0
                
                df['SGPA'] = df['SGPA'].apply(lambda x: f"{x:.2f}" if isinstance(x, (int, float)) and x != 0 else 'N/A')
                cols = base_cols + courses + end_cols
                df = df[cols]
                df = df.sort_values(by=['Student ID']).reset_index(drop=True)
                
                df.to_excel(writer, sheet_name=dept[:31], index=False, startrow=7)
                worksheet = writer.sheets[dept[:31]]
                
                max_col = len(cols)
                max_col_letter = get_column_letter(max_col)
                
                worksheet.merge_cells(f"A1:{max_col_letter}1")
                cell = worksheet["A1"]
                cell.value = "KTU RESULT ANALYSER - SUPPLEMENTARY"
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True, size=14)
                cell.alignment = center_align
                
                worksheet.merge_cells(f"A2:{max_col_letter}2")
                cell = worksheet["A2"]
                cell.value = f"{exam_name} ({detected_scheme} Scheme) ({detected_semester} Result)"
                cell.fill = green_fill
                cell.font = black_font
                cell.alignment = center_align
                
                worksheet.merge_cells(f"A3:{max_col_letter}3")
                cell = worksheet["A3"]
                cell.value = f"Department: {dept}"
                cell.fill = yellow_fill
                cell.font = black_font
                cell.alignment = center_align
                
                worksheet.merge_cells(f"A4:{max_col_letter}4")
                cell = worksheet["A4"]
                cell.value = college_name
                cell.fill = yellow_fill
                cell.font = black_font
                cell.alignment = center_align
                
                worksheet.merge_cells(f"A5:{max_col_letter}5")
                cell = worksheet["A5"]
                cell.value = f"Total Supplementary Students: {len(students)}"
                cell.fill = yellow_fill
                cell.font = black_font
                cell.alignment = center_align
                
                worksheet.merge_cells(f"A6:{max_col_letter}6")
                cell = worksheet["A6"]
                cell.value = "Note: This sheet contains Backlog/Supplementary students only."
                cell.font = red_font_italic
                cell.alignment = center_align
                
                for col_num in range(1, max_col + 1):
                    cell = worksheet.cell(row=8, column=col_num)
                    if cell.value == "Student ID": cell.value = "Register No"
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                fail_grades = ['F', 'FE', 'ABSENT', 'DEBARRED', 'I', 'WITHHELD']
                for row_num in range(9, 9 + len(df)):
                    for col_num in range(1, max_col + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.alignment = center_align
                        cell.border = thin_border
                        val = str(cell.value).strip().upper()
                        if val in fail_grades:
                            cell.font = red_font_bold
                        if col_num == max_col and str(cell.value) != '0' and str(cell.value) != 'N/A':
                            cell.font = red_font_bold
                        if col_num == 1:
                             cell.font = Font(color="000000", bold=True)
                             
                for col_num in range(1, max_col + 1):
                    col_letter = get_column_letter(col_num)
                    worksheet.column_dimensions[col_letter].width = 12
                worksheet.column_dimensions['A'].width = 18

        supp_buffer.seek(0)
        supp_base64 = base64.b64encode(supp_buffer.read()).decode('utf-8')

    overall_pass_percentage = round((total_passed / total_students) * 100, 2) if total_students > 0 else 0

    output = {
        "excelBase64": excel_base64,
        "stats": {
            "totalStudents": total_students,
            "passPercentage": overall_pass_percentage,
            "departments": stats_data
        }
    }
    if supp_base64:
        output["supplementaryExcelBase64"] = supp_base64
    if supp_total:
        output["supplementaryCount"] = supp_total
    if missing_credit_courses:
        output["missingCreditCourses"] = missing_credit_courses
    return output

@app.post("/api/convert")
async def convert_pdf(file: UploadFile = File(...)):
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="File must be a PDF")
        
    try:
        contents = await file.read()
        analysis_result = extract_and_analyze(contents)
        return analysis_result
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
